import re
import os
import json
import requests
import openpyxl
import time
import datetime
import random
from selenium import webdriver
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from concurrent.futures import ThreadPoolExecutor
from queue import Queue

current_datetime = datetime.datetime.now()
formatted_datetime = current_datetime.strftime("%Y-%m-%d_%H-%M-%S")

start_time = time.time()


def setup_driver():
    user_agent = UserAgent()

    options = webdriver.ChromeOptions()
    options.add_argument(f"user-agent={user_agent.random}")
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    return webdriver.Chrome(options=options)


# Функция для сохранения данных
def save_product_data(data, suffix):
    write_to_json(data, suffix)
    write_to_excel(data, suffix)


def process_product(link_to_the_product, product_data_queue):
    driver = setup_driver()
    try:
        pattern = r"\/product\/(\d+)\/"
        match = re.search(pattern, link_to_the_product)
        if not match:
            print("Product ID not found")
            return

        product_id = match.group(1)

        time.sleep(random.uniform(1, 3))
        driver.get(link_to_the_product)
        title = driver.title
        source = driver.page_source
        print(title)

        if "Zappos.com" not in title:
            print("Error retrieving page content.")
            return

        soup = BeautifulSoup(source, 'lxml')

        product_info = retrieve_product_info(soup, link_to_the_product)
        product_info["Product ID"] = product_id

        unique_image_urls = set()

        for picture_tag in soup.find_all("picture"):
            img_tag = picture_tag.find("img")
            if not img_tag:
                continue

            image_url = img_tag['src']
            if "SR920,736" not in image_url:
                modified_url = image_url.replace("_AC_SR73.60000000000001,58.88_", "920,736")
                unique_image_urls.add(modified_url)

        if not os.path.exists(product_info["Brand"]):
            os.mkdir(product_info["Brand"])

        dir_pic_name = []

        for url_p in unique_image_urls:
            img_data = requests.get(url=url_p, verify=False).content
            file_p_name = url_p.replace(".920,736", "").split("/")[-1]
            dir_pic_name.append(file_p_name)
            image_path = os.path.join(product_info["Brand"], file_p_name)
            with open(image_path, "wb") as handler:
                handler.write(img_data)

        product_info["Image Names"] = dir_pic_name
        product_data_queue.put(product_info)
    finally:
        driver.quit()


def following_links():
    list_item = open(file='zappos_shoes_urls.txt', mode='r', encoding='utf-8').read().splitlines()
    substring = "https://www.zappos.comhttps://zappos.ms.tagdelivery.com/click?id"
    filtered_list = [item for item in list_item if substring not in item]

    product_data_queue = Queue()

    saved_product_data = []

    # Фрагмент внутри цикла будет вызывать `save_product_data` после каждых n ссылок
    for i, link in enumerate(filtered_list):
        if (i + 1) % 1000 == 0:
            with ThreadPoolExecutor(max_workers=15) as executor:
                futures = [executor.submit(process_product, link, product_data_queue) for link in filtered_list[:i+1]]

            while not product_data_queue.empty():
                saved_product_data.append(product_data_queue.get())

            save_product_data(saved_product_data, f'partial_{i + 1}')
            saved_product_data = []

    # Запуск обработки оставшихся ссылок
    if filtered_list:
        with ThreadPoolExecutor(max_workers=15) as executor:
            futures = [executor.submit(process_product, link, product_data_queue) for link in filtered_list]

        while not product_data_queue.empty():
            saved_product_data.append(product_data_queue.get())

    return saved_product_data


def retrieve_product_info(soup, link_to_the_product=None):
    color_element = soup.find("span", class_="Tea-z")
    color = color_element.text if color_element else "Not found"

    brand_element = soup.find("span", itemprop="name")
    brand = brand_element.text.strip() if brand_element else "Not found"

    name_element = soup.find("span", class_="Sz-z")
    name = name_element.text.strip() if name_element else "Not found"

    gender_element = soup.find("span", class_="ksa-z").text.split(" ")[0]
    gender = gender_element if gender_element else "Not found"

    product_name = f'{brand} {name}'

    category_string = soup.find("div", id="breadcrumbs").text
    parts = category_string.strip().split('|')[1].split('/')
    categories = [part.strip() for part in parts]
    if len(categories) < 3:
        categories.append("")

    description_element = soup.find("div", itemprop="description")
    description = description_element.text.rstrip() if description_element else "Not found"

    text_without_view_zappos = description.split("View Zappos.com Glossary of Terms")[0].strip()

    clean_text = re.sub("SKU: #\d+", "", text_without_view_zappos).strip()

    product_info = {
        "Product Name": product_name,
        "Product Link": link_to_the_product,
        "Category 1": categories[0],
        "Category 2": categories[1],
        "Category 3": categories[2],
        "Color": color,
        "Brand": brand,
        "Gender": gender,
        "Product Information": clean_text
    }
    return product_info


def write_to_json(product_data, suffix):
    with open(f'zappos_data_{suffix}.json', 'w', encoding='utf-8') as f:
        json.dump(product_data, f, ensure_ascii=False, indent=4)


def write_to_excel(product_data, suffix):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = [
        'Product ID', 'Product Name', 'Product Link', 'Category 1', 'Category 2', 'Category 3', 'Color',
        'Gender', 'Brand', 'Product Information', 'Image Names'
    ]

    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    for row_num, product in enumerate(product_data, 2):
        for col_num, field in enumerate(headers, 1):
            if field == "Image Names":
                sheet.cell(row=row_num, column=col_num).value = ", ".join(product[field])
            else:
                sheet.cell(row=row_num, column=col_num).value = product[field]

    workbook.save(f"zappos_{suffix}.xlsx")


def main():
    product_data = following_links()
    save_product_data(product_data, 'final')


if __name__ == "__main__":
    main()

end_time = time.time()
total_time = round(end_time - start_time, 1)
print(f"Программа завершена за {total_time} секунд.")